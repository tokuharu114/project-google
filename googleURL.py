import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import NoSuchElementException
from selenium import webdriver              # Webブラウザを自動操作する（python -m pip install selenium)
                 # パスを通すためのコード
from csv import reader
import random
from selenium.webdriver.chrome.options import Options


import os
from openpyxl import Workbook


with open('kwlist.csv', 'r', encoding='utf-8') as csv_file:
    csv_reader = reader(csv_file)
    # Passing the cav_reader object to list() to get a list of lists
    list_of_rows = list(csv_reader)
    print(list_of_rows)


def GURL(list):
    options = Options()
    options.headless = False
    #driver = webdriver.Firefox()
    #driver = webdriver.Chrome(options=options)
    #options = Options()
    #options.add_argument('--headless')
    driver = webdriver.Chrome(executable_path="C:\\Users\\Administrator\\Downloads\\chromedriver.exe", options=options)                   # Chromeを準備

    # サンプルのHTMLを開く
    driver.get('https://www.google.com/')       # Googleを開く
    keywords=(list) 
    search = driver.find_element_by_name('q')   # HTML内で検索ボックス(name='q')を指定する
    search.send_keys(keywords)             # 検索ワードを送信する
    search.submit()                               # 3秒間待機
    time.sleep(random.randint(5,10))
    time.sleep(random.random())
    def ranking(driver):
        i = 1               # ループ番号、ページ番号を定義
        i_max = 11           # 最大何ページまで分析するかを定義
        title_list = []     # タイトルを格納する空リストを用意
        link_list = []      # URLを格納する空リストを用意
    
        # 現在のページが指定した最大分析ページを超えるまでループする
        while i <= i_max:
            print("page"+str(i)+" now getting url")
            # リンク('a href')とタイトル（'h3'）の入ったHTMLを抽出（クラス名で）→アプデで変わる可能性あり
            class_group = driver.find_elements_by_class_name('yuRUbf')
    
            # リンクとタイトルを抽出しリストに追加するforループ
            for elem in class_group:
    
                # リンク
                link = elem.find_element_by_tag_name('a').get_attribute('href')
    
                # タイトル
                title = elem.find_element_by_tag_name('h3').text
    
                # 動画、画像、説明, 「他の人は〜」といった記事とは関係ないコンテンツを排除
                if title != '動画' and title != '画像' and title != '説明':
                    if title != '':
                        title_list.append(title)
                        link_list.append(link)
    
            # 「次へ」は1つしかないが、あえてelementsで複数検索。空のリストであれば最終ページの意味になる。
            if driver.find_elements_by_id('pnnext') == []:
                i = i_max + 1
            else:
                # 次ページのURLはid="pnnext"のhref属性
                next_page = driver.find_element_by_id('pnnext').get_attribute('href')
    
                # 次ページへ遷移する
                driver.get(next_page)
                i = i + 1
                time.sleep(random.randint(3,8))
                time.sleep(random.random())
        return title_list, link_list
        
    

    # ranking関数を実行してタイトルとURLリストを取得する
    title, link = ranking(driver)
    print(ranking(driver))
    
    # タイトルリストをテキストに保存
    with open(str(keywords)+'_title.txt', mode='w', encoding='utf-8') as f:
        f.write("\n".join(title))
    
    # URLリストをテキストに保存
    with open(str(keywords)+'_link.csv', mode='w', encoding='utf-8') as f:
        f.write("\n".join(link))
    
    # ブラウザを閉じる
    driver.quit()

    

LEN=len(list_of_rows)
print(str(LEN))
for num in range(1,LEN):
    print("now getting url of "+str(list_of_rows[num]))
    
    GURL(list_of_rows[num]) 

filepathcsv="C:\DYM\getURL\\allcsv"
filepathtext='C:\DYM\getURL\\alltext'
basename = os.listdir (filepathcsv)
basenametext = os.listdir (filepathtext)
print(basename[0])
print(basenametext[0])
for i in range(len(basename)):
    with open('./alltext/'+basenametext[i], 'r', encoding='utf-8') as csv_file:
        csv_reader = reader(csv_file)
        # Passing the cav_reader object to list() to get a list of lists
        list_of_rows_text = list(csv_reader)
        #print(list_of_rows_text)

    with open('./allcsv/'+basename[i], 'r', encoding='utf-8') as csv_file:
        csv_reader = reader(csv_file)
        # Passing the cav_reader object to list() to get a list of lists
        list_of_rows_csv = list(csv_reader)
        #print(list_of_rows_csv)

    for p in range(len(list_of_rows_csv)):
        list_of_rows_csv[p].append(list_of_rows_text[p])


    fname=str(basename[i]).strip("['").strip("']_link.csv")
    # URLリストをテキストに保存
    #with open('./final/【'+fname+'_20230112】.csv', mode='w', encoding='utf-8') as f:
    #    f.write("\n".join(list_of_rows_csv))
    wb = Workbook()
    wb.save('./final/【'+fname+'_20230112】.xlsx')
    wb.create_sheet(fname)
    ws=wb[fname]
    ws = wb.active
        #print(list_of_rows_csv)
        #print(list_of_rows_csv[0][1])
        #print(list_of_rows_csv[0][0])
    for i in range(len(list_of_rows_csv)):
        ws.cell(row=1+i,column=1,value=list_of_rows_csv[i][0])
        ws.cell(row=1+i,column=2,value=str(list_of_rows_csv[i][1]).strip("['").strip("']_link.csv"))
    wb.save('./final/【'+fname+'_20230112】.xlsx')
    

